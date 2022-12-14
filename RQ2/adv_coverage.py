import shutil
import warnings
import sys
import time
from openpyxl import Workbook, load_workbook
warnings.filterwarnings("ignore")
import tensorflow as tf
config = tf.compat.v1.ConfigProto(gpu_options=tf.compat.v1.GPUOptions(allow_growth=True))
sess = tf.compat.v1.Session(config=config)
from keras import backend as K
import numpy as np
from coverage import Coverage
import os
#import pymannkendall as mk






def load_data(name):
    assert (name.upper() in ['MNIST', 'CIFAR', 'SVHN','FASHION_MNIST'])
    name = name.lower()
    x_train = np.load('../data/' + name + '_data/' + name + '_x_train.npy')
    y_train = np.load('../data/' + name + '_data/' + name + '_y_train.npy')
    x_test = np.load('../data/' + name + '_data/' + name + '_x_test.npy')
    y_test = np.load('../data/' + name + '_data/' + name + '_y_test.npy')
    return x_train, y_train, x_test, y_test

def retrain(model, X_train, Y_train, X_test, Y_test, batch_size=128, epochs=10):

    model.compile(
        loss='categorical_crossentropy',
        optimizer='adam',
        metrics=['accuracy']
    )

    # training without data augmentation
    history = model.fit(
        X_train, Y_train,
        epochs=epochs,
        batch_size=batch_size,
        shuffle=True,
        verbose=1,
        validation_data=(X_test, Y_test)
    )

    return model, history

if __name__ == '__main__':

    dataset = 'svhn'
    model_name = 'svhn_first'
    attack = 'cw-l2'

    num = 73257
    epoch = 15
    adv_epoch = [5]
    seed = 44
    R = 3

    x_train, y_train, x_test, y_test = load_data(dataset)
    x_adv = np.load('data/Adv_{0}_{1}_{2}.npy'.format(dataset, model_name, attack))
    y_adv = np.load('data/Adv_{0}_y.npy'.format(dataset))

    data_shape = [x_train.shape, x_adv.shape]
    print(data_shape)

    from sklearn.utils import shuffle

    x_train, y_train = shuffle(x_train, y_train, random_state=seed)
    x_adv, y_adv = shuffle(x_adv[:15000], y_adv[:15000], random_state=seed)

    x_mixed, y_mixed = np.vstack([x_train[:int(num * 0.8)], x_adv[:int(num * 0.2)]]), np.vstack(
        [y_train[:int(num * 0.8)], y_adv[:int(num * 0.2)]])
    x_mixed, y_mixed = shuffle(x_mixed, y_mixed, random_state=seed)
    print(x_mixed.shape, y_mixed.shape)

    # load mine trained model
    from keras.models import load_model

    model = load_model('../data/' + dataset + '_data/model/' + model_name + '.h5')
    model.summary

    model_layer = len(model.layers) - 1
    # data_shape = x_train.shape
    # print(data_shape)
    # print(model_layer)
    l = []
    model.summary()
    for i in range(1, model_layer):
        if model.layers[i].output.shape[1] != None:
            l.append(i)


    # ?????????????????????coverage list
    trend_train_nc1 = []
    trend_train_nc2 = []
    trend_train_nc3 = []
    trend_train_nc4 = []
    trend_train_nc5 = []
    trend_train_tknc = []
    trend_train_kmnc = []
    trend_train_tknp = []
    trend_train_nbc = []
    trend_train_snac = []
    trend_test_nc1 = []
    trend_test_nc2 = []
    trend_test_nc3 = []
    trend_test_nc4 = []
    trend_test_nc5 = []
    trend_test_tknc = []
    trend_test_kmnc = []
    trend_test_tknp = []
    trend_test_nbc = []
    trend_test_snac = []

    for i in range(epoch):
        print(i)
        print(time.asctime())
        model = load_model('new_model/' + dataset + '/'+ model_name +'/{}/model_{}.h5'.format(R, i))

        if i in adv_epoch:
            coverage = Coverage(model, x_mixed[0: num], x_mixed[0: num])
            nc1, activate_num1, total_num1 = coverage.NC(l, threshold=0.1)
            nc2, activate_num2, total_num2 = coverage.NC(l, threshold=0.3)
            nc3, activate_num3, total_num3 = coverage.NC(l, threshold=0.5)
            nc4, activate_num4, total_num4 = coverage.NC(l, threshold=0.7)
            nc5, activate_num5, total_num5 = coverage.NC(l, threshold=0.9)
            tknc, pattern_num, total_num6 = coverage.TKNC(l)
            tknp = coverage.TKNP(l)
            kmnc, nbc, snac, covered_num, l_covered_num, u_covered_num, neuron_num = coverage.KMNC(l)

            trend_train_nc1.append(nc1)
            trend_train_nc2.append(nc2)
            trend_train_nc3.append(nc3)
            trend_train_nc4.append(nc4)
            trend_train_nc5.append(nc5)
            trend_train_tknc.append(tknc)
            trend_train_kmnc.append(kmnc)
            trend_train_tknp.append(tknp)
            trend_train_nbc.append(nbc)
            trend_train_snac.append(snac)


            with open("coverage_result_of_train_data.txt", "a") as f:
                f.write("\n------------------------------------------------------------------------------\n")
                f.write('the result of {} {} adv_epoch{} experiment {} is: \n'.format(dataset, model_name, i, R))
                f.write('NC(0.1): {}  activate_num: {}  total_num: {} \n'.format(nc1, activate_num1, total_num1))
                f.write('NC(0.3): {}  activate_num: {}  total_num: {} \n'.format(nc2, activate_num2, total_num2))
                f.write('NC(0.5): {}  activate_num: {}  total_num: {} \n'.format(nc3, activate_num3, total_num3))
                f.write('NC(0.7): {}  activate_num: {}  total_num: {} \n'.format(nc4, activate_num4, total_num4))
                f.write('NC(0.9): {}  activate_num: {}  total_num: {} \n'.format(nc5, activate_num5, total_num5))
                f.write('TKNC: {}  pattern_num: {}  total_num: {} \n'.format(tknc, pattern_num, total_num6))
                f.write('TKNP: {} \n'.format(tknp))
                f.write('KMNC: {}  covered_num: {}  total_num: {} \n'.format(kmnc, covered_num, neuron_num))
                f.write('NBC: {}  l_covered_num: {}  u_covered_num: {} \n'.format(nbc, l_covered_num, u_covered_num))
                f.write('SNAC: {} \n'.format(snac))

        else:
            coverage = Coverage(model, x_train[0: num], x_train[0: num])
            nc1, activate_num1, total_num1 = coverage.NC(l, threshold=0.1)
            nc2, activate_num2, total_num2 = coverage.NC(l, threshold=0.3)
            nc3, activate_num3, total_num3 = coverage.NC(l, threshold=0.5)
            nc4, activate_num4, total_num4 = coverage.NC(l, threshold=0.7)
            nc5, activate_num5, total_num5 = coverage.NC(l, threshold=0.9)
            tknc, pattern_num, total_num6 = coverage.TKNC(l)
            tknp = coverage.TKNP(l)
            kmnc, nbc, snac, covered_num, l_covered_num, u_covered_num, neuron_num = coverage.KMNC(l)

            trend_train_nc1.append(nc1)
            trend_train_nc2.append(nc2)
            trend_train_nc3.append(nc3)
            trend_train_nc4.append(nc4)
            trend_train_nc5.append(nc5)
            trend_train_tknc.append(tknc)
            trend_train_kmnc.append(kmnc)
            trend_train_tknp.append(tknp)
            trend_train_nbc.append(nbc)
            trend_train_snac.append(snac)

            with open("coverage_result_of_train_data.txt", "a") as f:
                f.write("\n------------------------------------------------------------------------------\n")
                f.write('the result of {} {} epoch{} experiment {} is: \n'.format(dataset, model_name, i, R))
                f.write('NC(0.1): {}  activate_num: {}  total_num: {} \n'.format(nc1, activate_num1, total_num1))
                f.write('NC(0.3): {}  activate_num: {}  total_num: {} \n'.format(nc2, activate_num2, total_num2))
                f.write('NC(0.5): {}  activate_num: {}  total_num: {} \n'.format(nc3, activate_num3, total_num3))
                f.write('NC(0.7): {}  activate_num: {}  total_num: {} \n'.format(nc4, activate_num4, total_num4))
                f.write('NC(0.9): {}  activate_num: {}  total_num: {} \n'.format(nc5, activate_num5, total_num5))
                f.write('TKNC: {}  pattern_num: {}  total_num: {} \n'.format(tknc, pattern_num, total_num6))
                f.write('TKNP: {} \n'.format(tknp))
                f.write('KMNC: {}  covered_num: {}  total_num: {} \n'.format(kmnc, covered_num, neuron_num))
                f.write('NBC: {}  l_covered_num: {}  u_covered_num: {} \n'.format(nbc, l_covered_num, u_covered_num))
                f.write('SNAC: {} \n'.format(snac))




        # ??????test_data??????coverage
        """coverage = Coverage(model, x_train[0 : num], x_test[0 : num])
        nc1, activate_num1, total_num1 = coverage.NC(l, threshold=0.1)
        nc2, activate_num2, total_num2 = coverage.NC(l, threshold=0.3)
        nc3, activate_num3, total_num3 = coverage.NC(l, threshold=0.5)
        nc4, activate_num4, total_num4 = coverage.NC(l, threshold=0.7)
        nc5, activate_num5, total_num5 = coverage.NC(l, threshold=0.9)
        tknc, pattern_num, total_num6 = coverage.TKNC(l)
        tknp = coverage.TKNP(l)
        kmnc, nbc, snac, covered_num, l_covered_num, u_covered_num, neuron_num = coverage.KMNC(l)

        trend_test_nc1.append(nc1)
        trend_test_nc2.append(nc2)
        trend_test_nc3.append(nc3)
        trend_test_nc4.append(nc4)
        trend_test_nc5.append(nc5)
        trend_test_tknc.append(tknc)
        trend_test_kmnc.append(kmnc)
        trend_test_tknp.append(tknp)
        trend_test_nbc.append(nbc)
        trend_test_snac.append(snac)

        with open("coverage_result_of_test_data.txt", "a") as f:
            f.write("\n------------------------------------------------------------------------------\n")
            f.write('the result of {} {} epoch{} is: \n'.format(dataset, model_name, i))
            f.write('NC(0.1): {}  activate_num: {}  total_num: {} \n'.format(nc1, activate_num1, total_num1))
            f.write('NC(0.3): {}  activate_num: {}  total_num: {} \n'.format(nc2, activate_num2, total_num2))
            f.write('NC(0.5): {}  activate_num: {}  total_num: {} \n'.format(nc3, activate_num3, total_num3))
            f.write('NC(0.7): {}  activate_num: {}  total_num: {} \n'.format(nc4, activate_num4, total_num4))
            f.write('NC(0.9): {}  activate_num: {}  total_num: {} \n'.format(nc5, activate_num5, total_num5))
            f.write('TKNC: {}  pattern_num: {}  total_num: {} \n'.format(tknc, pattern_num, total_num6))
            f.write('TKNP: {} \n'.format(tknp))
            f.write('KMNC: {}  covered_num: {}  total_num: {} \n'.format(kmnc, covered_num, neuron_num))
            f.write('NBC: {}  l_covered_num: {}  u_covered_num: {} \n'.format(nbc, l_covered_num, u_covered_num))
            f.write('SNAC: {} \n'.format(snac))"""
        K.clear_session()


    train_trend = [['NC0.1'] + trend_train_nc1,
    ['NC0.3'] + trend_train_nc2,
    ['NC0.5'] + trend_train_nc3,
    ['NC0.7'] + trend_train_nc4,
    ['NC0.9'] + trend_train_nc5,
    ['TKNC'] + trend_train_tknc,
    ['TKNP'] + trend_train_tknp,
    ['KMNC'] + trend_train_kmnc,
    ['NBC'] + trend_train_nbc,
    ['SNAC'] + trend_train_snac]
    print('Coverage trend with {0} adv_examples in epoch {1}.xlsx'.format(attack, adv_epoch))
    if not os.path.isfile('Coverage trend with {0} adv_examples in epoch {1}.xlsx'.format(attack, adv_epoch)):
        wb = Workbook()
    else:
        wb = load_workbook('Coverage trend with {0} adv_examples in epoch {1}.xlsx'.format(attack, adv_epoch))
    ws = wb.create_sheet("train data of {0} {1} experiment {2}" .format(dataset, model_name, R))
    for row in train_trend:
        ws.append(row)
    wb.save('Coverage trend with {0} adv_examples in epoch {1}.xlsx'.format(attack, adv_epoch))

    """test_trend = [['NC0.1'] + trend_test_nc1,
    ['NC0.3'] + trend_test_nc2,
    ['NC0.5'] + trend_test_nc3,
    ['NC0.7'] + trend_test_nc4,
    ['NC0.9'] + trend_test_nc5,
    ['TKNC'] + trend_test_tknc,
    ['TKNP'] + trend_test_tknp,
    ['KMNC'] + trend_test_kmnc,
    ['NBC'] + trend_test_nbc,
    ['SNAC'] + trend_test_snac]

    ws = wb.create_sheet("test data of {0} {1}".format(dataset, model_name))
    for row in test_trend:
        ws.append(row)
    wb.save('Coverage trend with {0} adv_examples in epoch {1}.xlsx'.format(attack, adv_epoch))

    # print(trend_test_nc1,trend_train_snac)
    trend_result_train_nc1 = mk.original_test(trend_train_nc1)
    trend_result_train_nc2 = mk.original_test(trend_train_nc2)
    trend_result_train_nc3 = mk.original_test(trend_train_nc3)
    trend_result_train_nc4 = mk.original_test(trend_train_nc4)
    trend_result_train_nc5 = mk.original_test(trend_train_nc5)
    trend_result_train_tknc = mk.original_test(trend_train_tknc)
    trend_result_train_tknp = mk.original_test(trend_train_tknp)
    trend_result_train_kmnc = mk.original_test(trend_train_kmnc)
    trend_result_train_nbc = mk.original_test(trend_train_nbc)
    trend_result_train_snac = mk.original_test(trend_train_snac)
    with open("coverage_trend_of_train_data.txt", "a") as f:
        f.write("\n------------------------------------------------------------------------------\n")
        f.write("data number = {} \n".format(num))
        f.write('the result of {} {} experiment{} is: \n'.format(dataset, model_name, T))
        f.write('NC(0.1): {}  \n\n'.format(trend_result_train_nc1))
        f.write('NC(0.3): {}  \n\n'.format(trend_result_train_nc2))
        f.write('NC(0.5): {}  \n\n'.format(trend_result_train_nc3))
        f.write('NC(0.7): {}  \n\n'.format(trend_result_train_nc4))
        f.write('NC(0.9): {}  \n\n'.format(trend_result_train_nc5))
        f.write('TKNC: {}  \n\n'.format(trend_result_train_tknc))
        f.write('TKNP: {}  \n\n'.format(trend_result_train_tknp))
        f.write('KMNC: {}  \n\n'.format(trend_result_train_kmnc))
        f.write('NBC: {}   \n\n'.format(trend_result_train_nbc))
        f.write('SNAC: {}  \n\n'.format(trend_result_train_snac))

    trend_result_test_nc1 = mk.original_test(trend_test_nc1)
    trend_result_test_nc2 = mk.original_test(trend_test_nc2)
    trend_result_test_nc3 = mk.original_test(trend_test_nc3)
    trend_result_test_nc4 = mk.original_test(trend_test_nc4)
    trend_result_test_nc5 = mk.original_test(trend_test_nc5)
    trend_result_test_tknc = mk.original_test(trend_test_tknc)
    trend_result_test_tknp = mk.original_test(trend_test_tknp)
    trend_result_test_kmnc = mk.original_test(trend_test_kmnc)
    trend_result_test_nbc = mk.original_test(trend_test_nbc)
    trend_result_test_snac = mk.original_test(trend_test_snac)
    with open("coverage_trend_of_test_data.txt", "a") as f:
        f.write("\n------------------------------------------------------------------------------\n")
        f.write("data number = {} \n".format(num))
        f.write('the result of {} {} experiment{} is: \n'.format(dataset, model_name, T))
        f.write('NC(0.1): {}  \n\n'.format(trend_result_test_nc1))
        f.write('NC(0.3): {}  \n\n'.format(trend_result_test_nc2))
        f.write('NC(0.5): {}  \n\n'.format(trend_result_test_nc3))
        f.write('NC(0.7): {}  \n\n'.format(trend_result_test_nc4))
        f.write('NC(0.9): {}  \n\n'.format(trend_result_test_nc5))
        f.write('TKNC: {}  \n\n'.format(trend_result_test_tknc))
        f.write('TKNP: {}  \n\n'.format(trend_result_test_tknp))
        f.write('KMNC: {}  \n\n'.format(trend_result_test_kmnc))
        f.write('NBC: {}   \n\n'.format(trend_result_test_nbc))
        f.write('SNAC: {}  \n\n'.format(trend_result_test_snac))"""