from openpyxl import Workbook,load_workbook
import itertools
import numpy as np


if __name__ == '__main__':

    dataset = 'cifar'
    model_name = 'resnet20'
    attack = 'bim-b'

    num = 50000
    epoch = 15
    adv_epoch = [5]
    seed = 42
    R = 2

    wb = load_workbook('Coverage trend with {0} adv_examples in epoch {1}.xlsx'.format(attack, adv_epoch))
    ws = wb["train data {0} {1} R{2}".format(dataset, model_name, R)]
    nc1, nc2, nc3, nc4, nc5, tknc, tknp, kmnc, nbc, snac = [], [], [], [], [], [], [], [], [], []
    for (i, coverage) in ((1, nc1), (2, nc2), (3, nc3), (4, nc4), (5, nc5), (6, tknc), (7, tknp), (8, kmnc), (9,nbc), (10, snac)):
        cell_range = ws['B{0}'.format(i): 'P{0}'.format(i)]
        for each_cell in cell_range:
            for each in each_cell:
                coverage.append(each.value)
    #print(nc1, snac)

    nc1_ano, nc2_ano, nc3_ano, nc4_ano, nc5_ano, tknc_ano, tknp_ano, kmnc_ano, nbc_ano, snac_ano = ['NC0.1'], ['NC0.3'], ['NC0.5'], ['NC0.7'], ['NC0.9'], ['TKNC'], ['TKNP'], ['KMNC'], ['NBC'], ['SNAC']
    for (coverage,ano) in (nc1,nc1_ano), (nc2, nc2_ano), (nc3,nc3_ano), (nc4,nc4_ano), (nc5,nc5_ano), (tknc,tknc_ano), (tknp,tknp_ano), (kmnc, kmnc_ano), (nbc,nbc_ano), (snac,snac_ano):
        for j in range(15):
            if j == 0 or j == 14:
                ano.append(0)
            else:
                if coverage[j] > coverage[j-1] and coverage[j] > coverage[j+1]:
                    ano.append(1)
                elif coverage[j] < coverage[j-1] and coverage[j] < coverage[j+1]:
                    ano.append(1)
                else:
                    ano.append(0)
    #print( nc1_ano,"\n", snac_ano)
    for row in (nc1_ano, nc2_ano, nc3_ano, nc4_ano, nc5_ano, tknc_ano, tknp_ano, kmnc_ano, nbc_ano, snac_ano):
        ws.append(row)


    ano = (nc1_ano, nc2_ano, nc3_ano, nc4_ano, nc5_ano, tknc_ano, tknp_ano, kmnc_ano, nbc_ano, snac_ano)
    for num in range(1, 4):
        list = []

        for combinations in itertools.combinations(ano, num):
            # print(combinations)
            Px = 0
            Pxy = 0
            for i in range(1, 16):
                sum = 0
                for j in range(num):
                    if combinations[j][i] == 1:
                        sum = sum + 1
                if sum == num:
                    Px = Px + 1
            for i in adv_epoch:
                sum = 0
                for j in range(num):
                    if combinations[j][i+1] == 1:
                        sum = sum +1
                if sum == num:
                    Pxy = Pxy + 1
            # print(Px, Pxy)
            if Px != 0 and Pxy/Px >= 0.25:
                combinations = np.array(combinations)
                # print(combinations[:, 0])
                combinations = [Pxy/Px] + combinations[:, 0].tolist()
                list.append(combinations)
        # print(list)

        for row in list:
            ws.append(row)



    wb.save('Coverage trend with {0} adv_examples in epoch {1}.xlsx'.format(attack, adv_epoch))


