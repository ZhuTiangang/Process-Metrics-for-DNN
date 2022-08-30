import numpy as np
import pandas as pd
import matplotlib.pyplot as plt
from openpyxl import Workbook, load_workbook
from scipy.stats import spearmanr,pearsonr

if __name__ == '__main__':
    dataset = 'mnist'
    model_name = 'lenet5'

    num = 6000
    epoch = 30
    seed = 42
    R = 9

    wb = load_workbook('evaluate_' + dataset + '.xlsx')
    ws = wb['{0} {1} {2} {3}'.format(model_name, num, epoch, R)]
    acc, val_acc = [], []
    cell_range = ws['B3': 'AE3']
    for each_cell in cell_range:
        for each in each_cell:
            acc.append(each.value)
    cell_range = ws['B4': 'AE4']
    for each_cell in cell_range:
        for each in each_cell:
            val_acc.append(each.value)
    #print(acc,val_acc)
    overfitting = []
    for i in range(30):
        overfitting.append(acc[i] - val_acc[i])
    #print(overfitting)

    wb = load_workbook('Coverage of {0}.xlsx'.format(dataset))
    ws = wb["{0} {1} {2} {3}".format(model_name, num, epoch, R)]
    nc1, nc2, nc3, nc4, nc5, tknc, tknp, kmnc, nbc, snac = [], [], [], [], [], [], [], [], [], []
    for (i, coverage) in (
            (1, nc1), (2, nc2), (3, nc3), (4, nc4), (5, nc5), (6, tknc), (7, tknp), (8, kmnc), (9, nbc), (10, snac)):
        cell_range = ws['A{0}'.format(i): 'AE{0}'.format(i)]
        for each_cell in cell_range:
            for each in each_cell:
                coverage.append(each.value)

        spearman = spearmanr(coverage[1:], overfitting)
        pearson = pearsonr(coverage[1:], overfitting)
        print(coverage[0], spearman)
        print(pearson)
        with open("correlation_" + dataset + ".txt", "a") as f:
            f.write("\n------------------------------------------------------------------------------\n")
            f.write(model_name + '\t' + coverage[0] + "\t{}\n".format(R))
            f.write(str(spearman) + '\n')
            f.write('pearson_result ' + str(pearson))
